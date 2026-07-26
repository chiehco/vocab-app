"""xlsx -> JSON data pipeline for the vocab app.

Usage:
    python scripts/import_data.py --source <path-to-xlsx> [--out public/data/v1]

Reads the user's vocabulary spreadsheet (read-only) and writes versioned,
git-diffable JSON bundles consumed by the app at build time.
"""

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

VALID_LEVELS = {"LV1", "LV2", "LV3", "LV4", "LV5", "LV6"}
VALID_RELATION_TYPES = {"synonym", "antonym", "derivative", "word_form", "confuse", "root_family", "topic", "exam_distractor"}
VALID_RELATION_DIRECTIONS = {"one_way", "two_way"}
VALID_PRIORITY_TIERS = {"S", "A", "B", "C", "Z"}
VALID_SENSE_STATUSES = {"draft", "reviewed", "approved", "needs_check"}
VALID_HOOK_TYPES = {"phonetic", "association", "etymology"}
VALID_HOOK_SOURCE_TYPES = {"original", "adapted", "book_ref", "legacy"}
VALID_HOOK_SCOPES = {"private_only", "public_ok"}
VALID_HOOK_STATUSES = {"draft", "needs_check", "approved"}
HOOK_GUIDE_MARKER = "填表說明"

SHEETS = {
    "words": "input_words_單字主表",
    "senses": "input_senses_義項表",
    "examples": "input_examples_例句表",
    "relations": "input_relations_關聯詞",
    "morphemes": "input_morphemes_字根拆解",
    "media": "input_media_prompts_圖卡",
}

# 選配工作表：Excel 裡還沒有也不會報錯，輸出空陣列
OPTIONAL_SHEETS = {
    "notes": "input_notes_補充說明",
    "exam_priority": "exam_priority_學測優先",
    "hooks": "input_hooks_鉤子表",
}

ID_PREFIX = {"words": "W", "examples": "E", "relations": "R", "morphemes": "M", "media": "A", "notes": "N"}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def gen_id(sheet_key, *parts):
    h = hashlib.sha256(("|".join(str(p) for p in parts)).encode("utf-8")).hexdigest()[:8]
    return f"{ID_PREFIX[sheet_key]}X{h}"


def word_variants(entry):
    """Expand slash and parenthetical spellings while keeping the source entry intact."""
    variants = []
    for raw_part in entry.split("/"):
        part = raw_part.strip()
        match = re.fullmatch(r"([A-Za-z-]+)\(([A-Za-z-]+)\)", part)
        if not match:
            variants.append(part)
            continue
        base, addition = match.groups()
        common_prefix = 0
        for left, right in zip(base.casefold(), addition.casefold()):
            if left != right:
                break
            common_prefix += 1
        variants.extend([base, addition if common_prefix >= 3 else base + addition])
    return list(dict.fromkeys(v for v in variants if v))


def read_rows(ws, validate_instruction=True):
    """Yield cleaned row tuples starting from row 3 (row1=header, row2=instructional sample)."""
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    sample = next(rows, None)
    if header is None:
        return []
    # Row 2 of every input_* sheet is a 填表說明 row, not data. If the template
    # shape ever changes this assert fails loudly instead of silently dropping a word.
    if validate_instruction and sample is not None and clean(sample[0]) is not None:
        first = str(sample[0])
        if not ("可空" in first or "填" in first or "說明" in first):
            raise SystemExit(
                f"錯誤：工作表「{ws.title}」第 2 列看起來是真實資料（{first!r}），"
                "與預期的填表說明列不符。請檢查模板結構或調整 import_data.py。"
            )
    out = []
    for row in rows:
        vals = [clean(v) for v in row]
        if all(v is None for v in vals):
            continue
        out.append(vals)
    return out


def parse_words(ws):
    words = []
    seen = {}
    for r in read_rows(ws):
        (word_id, word, level, pos, meaning_zh, meaning_en, usage_pattern, syllables,
         stress, phonetic_us, family_key, is_core, source_note, status,
         image_word_id) = (list(r) + [None] * 15)[:15]
        if word is None:
            continue
        if word in seen:
            raise SystemExit(
                f"錯誤：單字「{word}」在 input_words 中重複出現（{seen[word]} 與 {word_id}）。"
                "進度資料以 word 為主鍵，必須唯一；請先在 Excel 中修正。"
            )
        seen[word] = word_id
        if level not in VALID_LEVELS:
            raise SystemExit(f"錯誤：單字「{word}」的 level 值「{level}」不在 LV1–LV6 範圍內。")
        pos = pos or ""
        pos_all = [p.strip() for p in pos.split("/") if p.strip()]
        record = {
            "wordId": word_id or gen_id("words", word),
            "word": word,
            "wordVariants": word_variants(word),
            "level": level,
            "pos": pos or None,
            "posAll": pos_all,
            "meaningZh": meaning_zh,
            "meaningEn": meaning_en,
            "usagePattern": usage_pattern,
            "syllables": syllables,
            "stressPattern": stress,
            "phoneticUs": phonetic_us,
            "familyKey": family_key,
            "isCore": bool(is_core) if is_core is not None else False,
            "sourceNote": source_note,
            "status": status or "draft",
        }
        if image_word_id:
            record["imageWordId"] = image_word_id
        words.append(record)
    words.sort(key=lambda w: w["wordId"])
    return words


def parse_examples(ws):
    out = []
    for r in read_rows(ws):
        (ex_id, word, sense_pos, hint, ex_type, sen_en, sen_zh, blank, answer,
         difficulty, status) = (list(r) + [None] * 11)[:11]
        if word is None or sen_en is None:
            continue
        out.append({
            "exampleId": ex_id or gen_id("examples", word, sen_en),
            "word": word,
            "sensePos": sense_pos,
            "meaningHint": hint,
            "exampleType": ex_type,
            "sentenceEn": sen_en,
            "sentenceZh": sen_zh,
            "blankSentence": blank,
            "answer": answer,
            "difficulty": difficulty,
            "status": status or "draft",
        })
    out.sort(key=lambda x: x["exampleId"])
    return out


def parse_senses(ws):
    """Read the sense sheet, whose guidance row is at the bottom instead of row 2."""
    rows = ws.iter_rows(min_row=1, values_only=True)
    next(rows, None)  # header
    out = []
    seen = set()
    for row_number, r in enumerate(rows, start=2):
        (sense_id, word_id, word, sense_pos, meaning_zh, is_exam_sense,
         exam_evidence, answer_forms, note, status) = (list(r) + [None] * 10)[:10]
        values = [clean(value) for value in (
            sense_id, word_id, word, sense_pos, meaning_zh, is_exam_sense,
            exam_evidence, answer_forms, note, status,
        )]
        (sense_id, word_id, word, sense_pos, meaning_zh, is_exam_sense,
         exam_evidence, answer_forms, note, status) = values
        if all(value is None for value in values):
            continue
        if not isinstance(sense_id, str) or not re.fullmatch(r"W\d{6}-\d+", sense_id):
            # The template keeps its field instructions in the last row.
            if isinstance(sense_id, str) and sense_id.startswith("義項唯一鍵"):
                continue
            raise SystemExit(f"錯誤：義項表第 {row_number} 列的 sense_id 無效：{sense_id!r}")
        if sense_id in seen:
            raise SystemExit(f"錯誤：義項表 sense_id 重複：{sense_id}")
        seen.add(sense_id)
        if not word_id or not word or not sense_pos or not meaning_zh:
            raise SystemExit(f"錯誤：義項表第 {row_number} 列缺少必要欄位：{sense_id}")
        if not sense_id.startswith(f"{word_id}-"):
            raise SystemExit(f"錯誤：義項 {sense_id} 與 word_id {word_id} 不一致")
        if is_exam_sense not in (None, "Y"):
            raise SystemExit(f"錯誤：義項 {sense_id} 的 is_exam_sense 必須為 Y 或留空")
        status = status or "draft"
        if status not in VALID_SENSE_STATUSES:
            raise SystemExit(f"錯誤：義項 {sense_id} 的 status 無效：{status}")
        forms = [] if answer_forms is None else [
            item.strip() for item in re.split(r"[/,;；、]+", str(answer_forms)) if item.strip()
        ]
        out.append({
            "senseId": sense_id,
            "wordId": word_id,
            "word": word,
            "senseOrder": int(sense_id.rsplit("-", 1)[1]),
            "sensePos": sense_pos,
            "meaningZh": meaning_zh,
            "isExamSense": is_exam_sense == "Y",
            "examEvidence": exam_evidence,
            "answerForms": forms,
            "note": note,
            "status": status,
        })
    out.sort(key=lambda row: (row["wordId"], row["senseOrder"]))
    return out


def parse_relations(ws):
    out = []
    seen = {}
    for row_number, r in enumerate(read_rows(ws), start=3):
        (rel_id, word, related, rel_type, direction, note, strength, status) = (list(r) + [None] * 8)[:8]
        if word is None or related is None:
            continue
        if word == related:
            raise SystemExit(f"錯誤：關聯詞第 {row_number} 列把「{word}」連到自己。")
        if rel_type not in VALID_RELATION_TYPES:
            raise SystemExit(f"錯誤：關聯詞第 {row_number} 列的 relation_type「{rel_type}」不合法。")
        if direction not in VALID_RELATION_DIRECTIONS:
            raise SystemExit(f"錯誤：關聯詞第 {row_number} 列的 direction「{direction}」不合法。")
        if not isinstance(strength, (int, float)) or not 1 <= strength <= 5:
            raise SystemExit(f"錯誤：關聯詞第 {row_number} 列的 strength 必須是 1–5。")
        key = (word.casefold(), related.casefold(), rel_type)
        if key in seen:
            print(
                f"警告：關聯詞第 {row_number} 列與第 {seen[key]} 列重複："
                f"{word} → {related}（{rel_type}），已保留第一筆。",
                file=sys.stderr,
            )
            continue
        seen[key] = row_number
        out.append({
            "relationId": rel_id or gen_id("relations", word, related, rel_type),
            "word": word,
            "relatedWord": related,
            "relationType": rel_type,
            "direction": direction,
            "note": note,
            "strength": strength,
            "status": status or "draft",
        })
    out.sort(key=lambda x: x["relationId"])
    return out


def parse_morphemes(ws):
    out = []
    for r in read_rows(ws):
        (row_id, word, morpheme, m_type, m_zh, m_en, origin, order, note, status) = (list(r) + [None] * 10)[:10]
        if word is None or morpheme is None:
            continue
        out.append({
            "rowId": row_id or gen_id("morphemes", word, morpheme, order),
            "word": word,
            "morpheme": morpheme,
            "morphemeType": m_type,
            "meaningZh": m_zh,
            "meaningEn": m_en,
            "origin": origin,
            "order": order,
            "note": note,
            "status": status or "draft",
        })
    out.sort(key=lambda x: x["rowId"])
    return out


def parse_media(ws):
    out = []
    for r in read_rows(ws):
        (asset_id, target_type, target_word, target_hint, media_type, image_type,
         prompt_en, caption_zh, status, license_note) = (list(r) + [None] * 10)[:10]
        if target_word is None or prompt_en is None:
            continue
        out.append({
            "assetId": asset_id or gen_id("media", target_word, prompt_en),
            "targetType": target_type,
            "targetWord": target_word,
            "targetHint": target_hint,
            "mediaType": media_type,
            "imageType": image_type,
            "promptEn": prompt_en,
            "captionZh": caption_zh,
            "status": status or "draft",
            "licenseNote": license_note,
        })
    out.sort(key=lambda x: x["assetId"])
    return out


def parse_exam_priority(ws):
    out = []
    for r in read_rows(ws, validate_instruction=False):
        values = (list(r) + [None] * 22)[:22]
        (rank, word_id, word, level, pos, meaning_zh, priority_tier, score_xuece,
         xt_base, xt_option, xt_cross, xt_years, xt_year_list, xt_n_option,
         xt_n_answer, adv_tier, score_zhikao, zk_years, zk_year_list,
         zk_n_option, zk_n_answer, is_function_word) = values
        if word is None or priority_tier not in VALID_PRIORITY_TIERS:
            continue
        out.append({
            "wordId": word_id,
            "word": word,
            "rank": rank,
            "level": level,
            "pos": pos,
            "meaningZh": meaning_zh,
            "priorityTier": priority_tier,
            "scoreXuece": score_xuece,
            "xtBase": xt_base,
            "xtOption": xt_option,
            "xtCross": xt_cross,
            "xtYears": xt_years,
            "xtYearList": xt_year_list,
            "xtOptionCount": xt_n_option,
            "xtAnswerCount": xt_n_answer,
            "advancedTier": adv_tier,
            "scoreZhikao": score_zhikao,
            "zkYears": zk_years,
            "zkYearList": zk_year_list,
            "zkOptionCount": zk_n_option,
            "zkAnswerCount": zk_n_answer,
            "isFunctionWord": str(is_function_word or "").upper() == "Y",
        })
    out.sort(key=lambda row: (row["rank"] if isinstance(row["rank"], (int, float)) else 10**9, row["word"]))
    return out


def parse_notes(ws):
    out = []
    for r in read_rows(ws):
        (note_id, word, note_type, title, content, status) = (list(r) + [None] * 6)[:6]
        if word is None or content is None:
            continue
        out.append({
            "noteId": note_id or gen_id("notes", word, note_type, title or content[:40]),
            "word": word,
            "noteType": note_type or "usage",
            "title": title,
            "content": content,
            "status": status or "draft",
        })
    out.sort(key=lambda x: x["noteId"])
    return out


def parse_hooks(ws):
    """鉤子表：義項表模式（第 2 列即資料、說明列在底部）。

    只跳過第一格等於 HOOK_GUIDE_MARKER 的說明列；其他格式錯誤一律報錯，
    不得默默當說明列跳過（規格 v3 §1）。
    """
    rows = ws.iter_rows(min_row=1, values_only=True)
    next(rows, None)  # header
    out = []
    seen = set()
    for row_number, r in enumerate(rows, start=2):
        values = [clean(v) for v in (list(r) + [None] * 13)[:13]]
        (hook_id, word_id, word, sense_id, target_meaning_zh, hook_type,
         hook_zh, hook_story, asset_id, source_type, source_ref,
         distribution_scope, status) = values
        if all(v is None for v in values):
            continue
        if hook_id == HOOK_GUIDE_MARKER:
            continue
        if not isinstance(hook_id, str) or not re.fullmatch(r"W\d{6}-H\d+", hook_id):
            raise SystemExit(f"錯誤：鉤子表第 {row_number} 列的 hook_id 無效：{hook_id!r}")
        if hook_id in seen:
            raise SystemExit(f"錯誤：鉤子表 hook_id 重複：{hook_id}")
        seen.add(hook_id)
        if not word_id or not hook_id.startswith(f"{word_id}-H"):
            raise SystemExit(f"錯誤：鉤子 {hook_id} 與 word_id {word_id!r} 不一致")
        if not word:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 缺少 word")
        if sense_id is None and target_meaning_zh is None:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 sense_id 與 target_meaning_zh 至少要填一個")
        if hook_type not in VALID_HOOK_TYPES:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 hook_type 無效：{hook_type!r}")
        if hook_type == "phonetic" and not hook_zh:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 為 phonetic 但 hook_zh 空白")
        if not hook_story:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 缺少 hook_story")
        if source_type not in VALID_HOOK_SOURCE_TYPES:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 source_type 無效：{source_type!r}")
        if source_type != "original" and not source_ref:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 source_type 為 {source_type}，source_ref 不可空白")
        if distribution_scope not in VALID_HOOK_SCOPES:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 distribution_scope 無效：{distribution_scope!r}")
        if status not in VALID_HOOK_STATUSES:
            raise SystemExit(f"錯誤：鉤子 {hook_id} 的 status 無效：{status!r}")
        out.append({
            "hookId": hook_id,
            "wordId": word_id,
            "word": word,
            "senseId": sense_id,
            "targetMeaningZh": target_meaning_zh,
            "hookType": hook_type,
            "hookZh": hook_zh,
            "hookStory": hook_story,
            "assetId": asset_id,
            "sourceType": source_type,
            "sourceRef": source_ref,
            "distributionScope": distribution_scope,
            "status": status,
        })
    out.sort(key=lambda x: x["hookId"])
    return out


def write_json(path, data):
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True)
    # 建置模式（規格 v3 §6，安全預設）：
    #   public（預設）＝鉤子只留 public_ok+approved，可寫入 public/
    #   private＝鉤子留所有 approved；qa＝鉤子全留（含未審）
    #   private/qa 的產物一律不得寫入 public/，防止私用書籍資料被 commit 進公開站
    ap.add_argument("--mode", choices=["public", "private", "qa"], default="public")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    src = Path(args.source)
    if not src.exists():
        raise SystemExit(f"錯誤：找不到來源檔案 {src}")
    default_out = {
        "public": "public/data/v1",
        "private": "dist-private/data/v1",
        "qa": "dist-qa/data/v1",
    }
    out_dir = Path(args.out if args.out is not None else default_out[args.mode])
    if args.mode != "public" and "public" in out_dir.parts:
        raise SystemExit(f"錯誤：{args.mode} 模式的產物不可寫入 public/ 目錄（{out_dir}）。")
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    for key, sheet_name in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"錯誤：找不到工作表「{sheet_name}」。")

    words = parse_words(wb[SHEETS["words"]])
    senses = parse_senses(wb[SHEETS["senses"]])
    examples = parse_examples(wb[SHEETS["examples"]])
    relations = parse_relations(wb[SHEETS["relations"]])
    morphemes = parse_morphemes(wb[SHEETS["morphemes"]])
    media = parse_media(wb[SHEETS["media"]])
    notes_sheet = OPTIONAL_SHEETS["notes"]
    notes = parse_notes(wb[notes_sheet]) if notes_sheet in wb.sheetnames else []
    priority_sheet = OPTIONAL_SHEETS["exam_priority"]
    exam_priority = parse_exam_priority(wb[priority_sheet]) if priority_sheet in wb.sheetnames else []
    hooks_sheet = OPTIONAL_SHEETS["hooks"]
    hooks = parse_hooks(wb[hooks_sheet]) if hooks_sheet in wb.sheetnames else []

    word_set = {w["word"] for w in words}
    word_ids = {w["wordId"]: w["word"] for w in words}
    mismatched_senses = [
        row for row in senses
        if row["wordId"] not in word_ids or word_ids[row["wordId"]] != row["word"]
    ]
    if mismatched_senses:
        sample = mismatched_senses[0]
        raise SystemExit(
            f"錯誤：義項表的 word_id 與單字主表不一致：{sample['senseId']} / "
            f"{sample['wordId']} / {sample['word']}"
        )
    missing_relation_words = sorted({
        endpoint
        for relation in relations
        for endpoint in (relation["word"], relation["relatedWord"])
        if endpoint not in word_set
    })
    if missing_relation_words:
        raise SystemExit(
            f"錯誤：關聯詞表中有 {len(missing_relation_words)} 個端點不存在於單字主表："
            + ", ".join(missing_relation_words[:10])
            + ("…" if len(missing_relation_words) > 10 else "")
        )
    for label, rows, field in [
        ("例句", examples, "word"),
        ("字根", morphemes, "word"),
        ("圖卡", media, "targetWord"),
        ("補充說明", notes, "word"),
        ("考試優先級", exam_priority, "word"),
    ]:
        missing = sorted({r[field] for r in rows if r[field] not in word_set})
        if missing:
            print(f"警告：{label}表中有 {len(missing)} 個單字不存在於單字主表：{', '.join(missing[:10])}"
                  + ("…" if len(missing) > 10 else ""), file=sys.stderr)
    mismatched_priority_ids = [
        row for row in exam_priority
        if row["wordId"] not in word_ids or word_ids[row["wordId"]] != row["word"]
    ]
    if mismatched_priority_ids:
        sample = mismatched_priority_ids[0]
        raise SystemExit(
            f"錯誤：考試優先級的 word_id 與主表不一致：{sample['wordId']} / {sample['word']}"
        )

    # 鉤子表跨表驗證（規格 v3 §6 驗證 2、3、5、7）
    senses_by_id = {row["senseId"]: row for row in senses}
    media_ids = {row["assetId"] for row in media}
    for hook in hooks:
        if hook["wordId"] not in word_ids or word_ids[hook["wordId"]] != hook["word"]:
            raise SystemExit(
                f"錯誤：鉤子 {hook['hookId']} 的 word_id/word 與主表不一致："
                f"{hook['wordId']} / {hook['word']}"
            )
        if hook["senseId"] is not None:
            sense = senses_by_id.get(hook["senseId"])
            if sense is None:
                raise SystemExit(f"錯誤：鉤子 {hook['hookId']} 的 sense_id 不存在：{hook['senseId']}")
            if sense["wordId"] != hook["wordId"]:
                raise SystemExit(
                    f"錯誤：鉤子 {hook['hookId']} 的 sense_id {hook['senseId']} 屬於另一個字（{sense['wordId']}）"
                )
            # 同填時以 sense_id 為準；中文必須等於義項 meaning_zh 或其「；」分段之一，避免兩邊漂移
            if hook["targetMeaningZh"] is not None:
                segments = [s.strip() for s in re.split(r"[；;]", sense["meaningZh"]) if s.strip()]
                if hook["targetMeaningZh"] != sense["meaningZh"] and hook["targetMeaningZh"] not in segments:
                    raise SystemExit(
                        f"錯誤：鉤子 {hook['hookId']} 的 target_meaning_zh「{hook['targetMeaningZh']}」"
                        f"與義項 {hook['senseId']} 的 meaning_zh「{sense['meaningZh']}」不一致"
                    )
        if hook["assetId"] is not None and hook["assetId"] not in media_ids:
            raise SystemExit(f"錯誤：鉤子 {hook['hookId']} 的 asset_id 不存在於圖卡表：{hook['assetId']}")

    write_json(out_dir / "words.json", words)
    write_json(out_dir / "senses.json", senses)
    write_json(out_dir / "examples.json", examples)
    write_json(out_dir / "relations.json", relations)
    write_json(out_dir / "morphemes.json", morphemes)
    write_json(out_dir / "media.json", media)
    write_json(out_dir / "notes.json", notes)
    write_json(out_dir / "exam_priority.json", exam_priority)

    # 依建置模式過濾鉤子（安全預設：public 只出 public_ok+approved）
    if args.mode == "public":
        hooks_out = [h for h in hooks
                     if h["distributionScope"] == "public_ok" and h["status"] == "approved"]
    elif args.mode == "private":
        hooks_out = [h for h in hooks if h["status"] == "approved"]
    else:  # qa：唯一保留 draft／needs_check 的模式
        hooks_out = hooks
    write_json(out_dir / "hooks.json", hooks_out)

    words_hash = hashlib.sha256((out_dir / "words.json").read_bytes()).hexdigest()
    # contentHash 涵蓋所有 App 會載入的資料檔：任何一張表變動都會觸發前端重灌
    h = hashlib.sha256()
    for name in ["words.json", "senses.json", "examples.json", "relations.json", "morphemes.json", "notes.json", "exam_priority.json", "hooks.json"]:
        h.update((out_dir / name).read_bytes())
    content_hash = h.hexdigest()
    meta = {
        "schemaVersion": 2,
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "sourceFile": src.name,
        "counts": {
            "words": len(words),
            "senses": len(senses),
            "examples": len(examples),
            "relations": len(relations),
            "morphemes": len(morphemes),
            "media": len(media),
            "notes": len(notes),
            "examPriority": len(exam_priority),
            "hooks": len(hooks_out),
        },
        "wordsHash": words_hash,
        "contentHash": content_hash,
    }
    write_json(out_dir / "meta.json", meta)

    print(f"匯入完成（mode={args.mode} → {out_dir}）：")
    for k, v in meta["counts"].items():
        print(f"  {k}: {v}")
    print(f"  wordsHash: {words_hash[:16]}…")


if __name__ == "__main__":
    main()
